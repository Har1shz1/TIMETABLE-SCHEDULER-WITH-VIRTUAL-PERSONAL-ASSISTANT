import random
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill 
import csv
  
class Faculty:
    def _init_(self, faculty_id, name, gender, courses, priority=None):
        self.id = faculty_id
        self.name = name
        self.gender = gender
        self.courses = list(courses)
        self.classes_allotted = []
        self.priority = priority


class Course:
    def _init_(self, code, course_title, credits, theory_slots, lab_slots=None):
        self.code = code
        self.course_title = course_title
        self.credits = credits
        self.no_theory_slots = theory_slots
        self.no_lab_slots = lab_slots


# class Room:
#     def _init_(self, rooms):
#         self.rooms = list(rooms)


class SlotInfo:
    def _init_(self):
        self.slot = ''
        self.allotted_theory_classes = {
            "course": [],
            "faculty": [],
            "room": []
        }
        self.allotted_lab_classes = {
            "course": [],
            "slot": [],
            "faculty": [],
            "room": []
        }

    def add_theory_class(self, course, faculty, room):
        self.allotted_theory_classes["course"].append(course)
        self.allotted_theory_classes["faculty"].append(faculty)
        self.allotted_theory_classes["room"].append(room)

    def add_lab_class(self, course, faculty, room):
        self.allotted_lab_classes["course"].append(course)
        self.allotted_lab_classes["faculty"].append(faculty)
        self.allotted_lab_classes["room"].append(room)


# Lab slot implementation is remaining.
# Implemented only got 4 Credit Theory Courses
class Timetable:
    def _init_(self):
        self.main_timetable = [
            [SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(),          0, SlotInfo(), 0, SlotInfo(), -1, SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(), SlotInfo()],
            [SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(),          0, SlotInfo(), 0, SlotInfo(), -1, SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(),          0, SlotInfo(), SlotInfo()],
            [SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(),          0, SlotInfo(), 0, SlotInfo(), -1, SlotInfo(),          0, SlotInfo(), SlotInfo(),          0, SlotInfo(),          0, SlotInfo(), SlotInfo()],
            [SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(), 0, SlotInfo(), -1, SlotInfo(),          0, SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo(), SlotInfo()],
            [SlotInfo(), SlotInfo(), SlotInfo(),          0, SlotInfo(), SlotInfo(), SlotInfo(), 0, SlotInfo(), -1, SlotInfo(),          0, SlotInfo(), SlotInfo(),          0, SlotInfo(),          0, SlotInfo(), SlotInfo()]
        ]

        self.days = ['TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
        self.days_slot = {
            "Tuesday": ['L1+L2', 'L2+L3', 'L3+L4', 'L4+L5', 'L5+L6', 'L31+L32', 'L32+L33', 'L33+L34', 'L34+L35', 'L35+L36'],
            "Wednesday": ['L7+L8', 'L8+L9', 'L9+L10', 'L10+L11', 'L11+L12', 'L37+L38', 'L38+L39', 'L39+L40', 'L40+L41', 'L41+L42'],
            "Thursday": ['L13+L14', 'L14+L15', 'L15+L16', 'L16+L17', 'L17+L18', 'L43+L44', 'L44+L45', 'L45+L46', 'L46+L47', 'L47+L48'],
            "Friday": ['L19+L20', 'L20+L21', 'L21+L22', 'L22+L23', 'L23+L24', 'L49+L50', 'L50+L51', 'L51+L52', 'L52+L53', 'L53+L54'],
            "Saturday": ['L25+L26', 'L26+L27', 'L27+L28', 'L28+L29', 'L29+L30', 'L55+L56', 'L56+L57', 'L57+L58', 'L58+L59', 'L59+L60']
        }

        # Make the slots more detailed, i.e. add TA1, TAA1, etc. separately.
        # Add SE1, SD1, SF1 slots location too
        self.theory_slots = {
            'A1': [self.main_timetable[0][1], self.main_timetable[2][4], self.main_timetable[3][2], self.main_timetable[4][2]],
            'B1': [self.main_timetable[0][2], self.main_timetable[1][6], self.main_timetable[3][1], self.main_timetable[4][4]],
            'C1': [self.main_timetable[0][4], self.main_timetable[2][4], self.main_timetable[3][0], self.main_timetable[4][1]],
            'D1': [self.main_timetable[0][6], self.main_timetable[1][1], self.main_timetable[2][2], self.main_timetable[4][0]],
            'E1': [self.main_timetable[0][0], self.main_timetable[1][4], self.main_timetable[3][4], self.main_timetable[4][6]],
            'F1': [self.main_timetable[1][2], self.main_timetable[2][0], self.main_timetable[2][6], self.main_timetable[3][6]],
            # 'G1': [self.main_timetable[1][0], self.main_timetable[3][3], self.main_timetable[4][5]],
            'A2': [self.main_timetable[0][12], self.main_timetable[2][15], self.main_timetable[3][13], self.main_timetable[4][13]],
            'B2': [self.main_timetable[0][13], self.main_timetable[1][15], self.main_timetable[2][10], self.main_timetable[3][12]],
            'C2': [self.main_timetable[0][15], self.main_timetable[1][17], self.main_timetable[3][10], self.main_timetable[4][12]],
            'D2': [self.main_timetable[0][17], self.main_timetable[1][12], self.main_timetable[2][13], self.main_timetable[4][10]],
            'E2': [self.main_timetable[0][10], self.main_timetable[1][10], self.main_timetable[3][15], self.main_timetable[4][17]]
            # 'F2': [self.main_timetable[1][13], self.main_timetable[2][12], self.main_timetable[3][17]],
            # 'G2': [self.main_timetable[0][14], self.main_timetable[2][17], self.main_timetable[3][14]]
        }

        self.lab_slots = {
            'L1+L2': [self.main_timetable[0][0], self.main_timetable[0][1]],
            'L2+L3': [self.main_timetable[0][1], self.main_timetable[0][2]],
            'L3+L4': [self.main_timetable[0][2], self.main_timetable[0][4]],
            'L4+L5': [self.main_timetable[0][4], self.main_timetable[0][6]],
            'L5+L6': [self.main_timetable[0][6], self.main_timetable[0][8]],
            'L7+L8': [self.main_timetable[1][0], self.main_timetable[1][1]],
            'L8+L9': [self.main_timetable[1][1], self.main_timetable[1][2]],
            'L9+L10': [self.main_timetable[1][2], self.main_timetable[1][4]],
            'L10+L11': [self.main_timetable[1][4], self.main_timetable[1][6]],
            'L11+L12': [self.main_timetable[1][6], self.main_timetable[1][8]],
            'L13+L14': [self.main_timetable[2][0], self.main_timetable[2][1]],
            'L14+L15': [self.main_timetable[2][1], self.main_timetable[2][2]],
            'L15+L16': [self.main_timetable[2][2], self.main_timetable[2][4]],
            'L16+L17': [self.main_timetable[2][4], self.main_timetable[2][6]],
            'L17+L18': [self.main_timetable[2][6], self.main_timetable[2][8]],
            'L19+L20': [self.main_timetable[3][0], self.main_timetable[3][1]],
            'L20+L21': [self.main_timetable[3][1], self.main_timetable[3][2]],
            'L21+L22': [self.main_timetable[3][2], self.main_timetable[3][4]],
            'L22+L23': [self.main_timetable[3][4], self.main_timetable[3][6]],
            'L23+L24': [self.main_timetable[3][6], self.main_timetable[3][8]],
            'L25+L26': [self.main_timetable[4][0], self.main_timetable[4][1]],
            'L26+L27': [self.main_timetable[4][1], self.main_timetable[4][2]],
            'L27+L28': [self.main_timetable[4][2], self.main_timetable[4][4]],
            'L28+L29': [self.main_timetable[4][4], self.main_timetable[4][6]],
            'L29+L30': [self.main_timetable[4][6], self.main_timetable[4][8]],
            # Afternoon slots
            'L31+L32': [self.main_timetable[0][10], self.main_timetable[0][12]],
            'L32+L33': [self.main_timetable[0][12], self.main_timetable[0][13]],
            'L33+L34': [self.main_timetable[0][13], self.main_timetable[0][15]],
            'L34+L35': [self.main_timetable[0][15], self.main_timetable[0][17]],
            'L35+L36': [self.main_timetable[0][17], self.main_timetable[0][18]],
            'L37+L38': [self.main_timetable[1][10], self.main_timetable[1][12]],
            'L38+L39': [self.main_timetable[1][12], self.main_timetable[1][13]],
            'L39+L40': [self.main_timetable[1][13], self.main_timetable[1][15]],
            'L40+L41': [self.main_timetable[1][15], self.main_timetable[1][17]],
            'L41+L42': [self.main_timetable[1][17], self.main_timetable[1][18]],
            'L43+L44': [self.main_timetable[2][10], self.main_timetable[2][12]],
            'L44+L45': [self.main_timetable[2][12], self.main_timetable[2][13]],
            'L45+L46': [self.main_timetable[2][13], self.main_timetable[2][15]],
            'L46+L47': [self.main_timetable[2][15], self.main_timetable[2][17]],
            'L47+L48': [self.main_timetable[2][17], self.main_timetable[2][18]],
            'L49+L50': [self.main_timetable[3][10], self.main_timetable[3][12]],
            'L50+L51': [self.main_timetable[3][12], self.main_timetable[3][13]],
            'L51+L52': [self.main_timetable[3][13], self.main_timetable[3][15]],
            'L52+L53': [self.main_timetable[3][15], self.main_timetable[3][17]],
            'L53+L54': [self.main_timetable[3][17], self.main_timetable[3][18]],
            'L55+L56': [self.main_timetable[4][10], self.main_timetable[4][12]],
            'L56+L57': [self.main_timetable[4][12], self.main_timetable[4][13]],
            'L57+L58': [self.main_timetable[4][13], self.main_timetable[4][15]],
            'L58+L59': [self.main_timetable[4][15], self.main_timetable[4][17]],
            'L59+L60': [self.main_timetable[4][17], self.main_timetable[4][18]]
        }

    def schedule_timetable_theory(self, faculty_list, rooms_list):
        for faculty in faculty_list:
            print(f'Scheduling for {faculty.name}.')

            available_slots = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'A2', 'B2', 'C2', 'D2', 'E2']  # Removed G1, F2, G2
            for course in faculty.courses:
                print(f'Allotting slot for "{course.course_title}".')

                # Add more if else statements for other types of courses.
                if course.credits == 4 and course.no_theory_slots == 3 and course.no_lab_slots == 1:
                    flag = True
                    while flag:
                        random_slot = random.choice(available_slots)
                        random_room = random.choice(rooms_list)
                        if random_room in self.theory_slots[random_slot][0].allotted_theory_classes["room"]:
                            continue
                        else:
                            for i, slot in enumerate(self.theory_slots[random_slot]):
                                # Temp condition for allotting only 3 slots for theory as there are 4 slots available total.
                                if i == 3:
                                    break
                                slot.allotted_theory_classes["course"].append(course)
                                slot.allotted_theory_classes["faculty"].append(faculty)
                                slot.allotted_theory_classes["room"].append(random_room)
                            available_slots.remove(random_slot)
                            flag = False
                print(f'Slot allotted for "{course.course_title}".')
            print(f'Timetable prepared for {faculty.name}.')
            print(f'--------------------------------------')
        print("All timetables scheduled successfully.")
        self.save_timetable_theory(faculty_list)

    # FIX REMOVING AVAILABLE SLOTS
    def schedule_research_scholar_lab_timetable(self, faculty_list, rooms_list, slots_offered=None):
        for faculty in faculty_list:
            print(f'Scheduling for {faculty.name}.')

            # IDEA: Make this a dictionary and introduce days here itself.
            # available_slots = slots_offered
            available_slots = ['L1+L2', 'L2+L3', 'L3+L4', 'L4+L5', 'L5+L6', 'L7+L8', 'L8+L9', 'L9+L10', 'L10+L11',
                               'L11+L12', 'L13+L14', 'L14+L15', 'L15+L16', 'L16+L17', 'L17+L18', 'L19+L20', 'L20+L21',
                               'L21+L22', 'L22+L23', 'L23+L24', 'L25+L26', 'L26+L27', 'L27+L28', 'L28+L29', 'L29+L30',
                               'L31+L32', 'L32+L33', 'L33+L34', 'L34+L35', 'L35+L36', 'L37+L38', 'L38+L39', 'L39+L40',
                               'L40+L41', 'L41+L42', 'L43+L44', 'L44+L45', 'L45+L46', 'L46+L47', 'L47+L48', 'L49+L50',
                               'L50+L51', 'L51+L52', 'L52+L53', 'L53+L54', 'L55+L56', 'L56+L57', 'L57+L58', 'L58+L59',
                               'L59+L60']

            assigned_slots = {
                "Tuesday": [],
                "Wednesday": [],
                "Thursday": [],
                "Friday": [],
                "Saturday": [],
            }

            available_days = [0, 0, 0, 0, 0]  # Tuesday, Wednesday, Thursday, Friday, Saturday | Tracks the number of labs in a day.
            early_morning_slots = ['L1+L2', 'L7+L8', 'L13+L14', 'L19+L20', 'L25+L26']
            late_evening_slots = ['L35+L36', 'L41+L42', 'L47+L48', 'L53+L54', 'L59+L60']

            # Condition for no female faculty should get a class after 5.
            if faculty.gender == "Female":
                late = ['L34+L35', 'L35+L36', 'L40+L41', 'L41+L42', 'L46+L47', 'L47+L48', 'L52+L53', 'L53+L54', 'L58+L59', 'L59+L60']
                available_slots = [slot for slot in available_slots if slot not in late]

            for course in faculty.courses:
                print(f'Allotting slot for "{course.course_title}".')

                if course.no_lab_slots == 1:
                    flag = True
                    while flag:
                        random_slot = random.choice(available_slots)
                        random_room = random.choice(rooms_list)
                        day = str(self.find_key(self.days_slot, random_slot))

                        # Condition for no more than 2 labs in a day.
                        if available_days[self.days.index(day.upper())] > 1:
                            continue

                        if random_room in self.lab_slots[random_slot][0].allotted_lab_classes["room"]:
                            continue
                        # Only 1 early morning or late evening slot in a week. If one or the other is chosen from
                        # random_slot other slots in early or late is removed. So it also enforces the 8 - 5 or 9 - 7:30
                        # constraint.
                        elif random_slot in (early_morning_slots or late_evening_slots):
                            for i, slot in enumerate(self.lab_slots[random_slot]):
                                if i == 2:
                                    break
                                slot.allotted_lab_classes["course"].append(course)
                                slot.allotted_lab_classes["faculty"].append(faculty)
                                slot.allotted_lab_classes["room"].append(random_room)
                                slot.allotted_lab_classes["slot"].append(random_slot)
                            available_slots.remove(random_slot)
                            available_slots = [slot for slot in available_slots if slot not in early_morning_slots]
                            available_slots = [slot for slot in available_slots if slot not in late_evening_slots]

                            available_days[self.days.index(day.upper())] += 1
                            assigned_slots[day].append(random_slot)

                            # Removes the slots with half clash and continuous slots
                            for slot in available_slots:
                                if self.are_slots_continuous(slot, random_slot) or self.are_slots_half_clash(slot, random_slot):
                                    available_slots.remove(slot)

                            flag = False
                        else:
                            for i, slot in enumerate(self.lab_slots[random_slot]):
                                if i == 2:
                                    break
                                slot.allotted_lab_classes["course"].append(course)
                                slot.allotted_lab_classes["faculty"].append(faculty)
                                slot.allotted_lab_classes["room"].append(random_room)
                                slot.allotted_lab_classes["slot"].append(random_slot)

                            available_slots.remove(random_slot)
                            available_days[self.days.index(day.upper())] += 1
                            assigned_slots[day].append(random_slot)

                            # Removes the slots with half clash and continuous slots
                            for slot in available_slots:
                                if self.are_slots_continuous(slot, random_slot) or self.are_slots_half_clash(slot, random_slot):
                                    available_slots.remove(slot)

                            flag = False
                print(f'Slot allotted for "{course.course_title}".')
            print(f'Timetable prepared for {faculty.name}.')
            print(f'--------------------------------------')
        print("All timetables scheduled successfully.")
        self.save_timetable_lab(faculty_list)

    def save_timetable_theory(self, faculty_list):
        wb = openpyxl.Workbook()
        for faculty in faculty_list:
            sheet = wb.create_sheet(title=faculty.name)

            days = ['TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
            timings = ['8:00\n8:50', '9:00\n9:50', '10:00\n10:50', '10:00\n10:50', '11:00\n11:50', '11:00\n11:50', '12:00\n12:50', '12:00\n12:50', '13:00\n13:50', 'LUNCH', '14:00\n14:50', '14:00\n14:50', '15:00\n15:50', '16:00\n16:50', '16:00\n16:50', '17:00\n17:50', '17:00\n17:50', '18:00\n18:50', '19:00\n19:50']
            for i in range(5):
                cell = sheet.cell(row=i+3, column=2)
                cell.value = days[i]
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFE2E3E3", end_color="FFE2E3E3", fill_type="solid")

            for i in range(19):
                cell = sheet.cell(row=2, column=i + 3)
                cell.value = timings[i]
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFCCCCFF", end_color="FFCCCCFF", fill_type="solid")

            for i in range(5):
                for j in range(19):
                    if self.main_timetable[i][j] == 0:
                        cell = sheet.cell(row=i + 3, column=j + 3)
                        cell.value = "-"
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
                    elif self.main_timetable[i][j] == -1:
                        cell = sheet.cell(row=i + 3, column=j + 3)
                        cell.value = "LUNCH"
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFE2E3E3", end_color="FFE2E3E3", fill_type="solid")
                    else:
                        if faculty in self.main_timetable[i][j].allotted_theory_classes["faculty"]:
                            index = self.main_timetable[i][j].allotted_theory_classes["faculty"].index(faculty)
                            cell = sheet.cell(row=i + 3, column=j + 3)
                            cell.value = f'{self.find_key(self.theory_slots, self.main_timetable[i][j])} \n{self.main_timetable[i][j].allotted_theory_classes["course"][index].code} \n{self.main_timetable[i][j].allotted_theory_classes["room"][index]}'
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color="FFCCFE32", end_color="FFCCFE32", fill_type="solid")
                        else:
                            cell = sheet.cell(row=i + 3, column=j + 3)
                            cell.value = "-"
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin', color='FF3D8CBC'),
                right=Side(style='thin', color='FF3D8CBC'),
                top=Side(style='thin', color='FF3D8CBC'),
                bottom=Side(style='thin', color='FF3D8CBC')
            )

            for row in sheet.iter_rows(min_row=2, max_row=7, min_col=2, max_col=21):
                for cell in row:
                    cell.border = thin_border

            wb.save("Timetable.xlsx")

    def save_timetable_lab(self, faculty_list):
        wb = openpyxl.Workbook()
        for faculty in faculty_list:
            sheet = wb.create_sheet(title=faculty.name)

            timings = ['8:00\n8:50', '9:00\n9:50', '10:00\n10:50', '10:00\n10:50', '11:00\n11:50', '11:00\n11:50', '12:00\n12:50', '12:00\n12:50', '13:00\n13:50', 'LUNCH', '14:00\n14:50', '14:00\n14:50', '15:00\n15:50', '16:00\n16:50', '16:00\n16:50', '17:00\n17:50', '17:00\n17:50', '18:00\n18:50', '19:00\n19:50']
            for i in range(5):
                cell = sheet.cell(row=i+3, column=2)
                cell.value = self.days[i]
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFE2E3E3", end_color="FFE2E3E3", fill_type="solid")

            for i in range(19):
                cell = sheet.cell(row=2, column=i + 3)
                cell.value = timings[i]
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFCCCCFF", end_color="FFCCCCFF", fill_type="solid")

            for i in range(5):
                for j in range(19):
                    if self.main_timetable[i][j] == 0:
                        cell = sheet.cell(row=i + 3, column=j + 3)
                        cell.value = "-"
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
                    elif self.main_timetable[i][j] == -1:
                        cell = sheet.cell(row=i + 3, column=j + 3)
                        cell.value = "LUNCH"
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFE2E3E3", end_color="FFE2E3E3", fill_type="solid")
                    else:
                        if faculty in self.main_timetable[i][j].allotted_lab_classes["faculty"]:
                            index = self.main_timetable[i][j].allotted_lab_classes["faculty"].index(faculty)
                            cell = sheet.cell(row=i + 3, column=j + 3)
                            cell.value = f'{self.main_timetable[i][j].allotted_lab_classes["slot"][index]} \n{self.main_timetable[i][j].allotted_lab_classes["course"][index].code} \n{self.main_timetable[i][j].allotted_lab_classes["room"][index]}'
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color="FFCCFE32", end_color="FFCCFE32", fill_type="solid")
                        else:
                            cell = sheet.cell(row=i + 3, column=j + 3)
                            cell.value = "-"
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin', color='FF3D8CBC'),
                right=Side(style='thin', color='FF3D8CBC'),
                top=Side(style='thin', color='FF3D8CBC'),
                bottom=Side(style='thin', color='FF3D8CBC')
            )

            for row in sheet.iter_rows(min_row=2, max_row=7, min_col=2, max_col=21):
                for cell in row:
                    cell.border = thin_border

            wb.save("Timetable.xlsx")

    @staticmethod
    def find_key(d, value):
        for key, values in d.items():
            if value in values:
                return key
        return None

    @staticmethod
    def find_row(timetable1, element):
        for row_index, row in enumerate(timetable1):
            if element in row:
                return row_index
        return -1

    @staticmethod
    def are_slots_continuous(slot1, slot2):
        slot1_end = int(re.findall(r'\d+', slot1.split('+')[-1])[0]) # Get the last slot number from slot1
        slot2_start = int(re.findall(r'\d+', slot2.split('+')[0])[0]) # Get the first slot number from slot2
        return slot1_end + 1 == slot2_start

    @staticmethod
    def are_slots_half_clash(slot1, slot2):
        # Extract all slot numbers using a regular expression
        slot1_start = int(re.findall(r'\d+', slot1.split('+')[0])[0])  # Get the first slot number from slot1
        slot1_end = int(re.findall(r'\d+', slot1.split('+')[-1])[0])  # Get the last slot number from slot1
        slot2_start = int(re.findall(r'\d+', slot2.split('+')[0])[0])  # Get the first slot number from slot2
        slot2_end = int(re.findall(r'\d+', slot2.split('+')[-1])[0])  # Get the last slot number from slot2
        return slot1_end == slot2_start or slot1_start == slot2_end


def load_courses_from_csv(file_path):
    courses = []
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            course_id = row['course_id']
            name = row['course_name']
            credits = int(row['credits'])
            theory_slots = int(row['theory_slots'])
            lab_slots = int(row['lab_slots'])
            course = Course(course_id, name, credits, theory_slots, lab_slots)
            courses.append(course)
    return courses


def load_faculties_from_csv(file_path, courses):
    faculties = []
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            faculty_id = int(row['faculty_id'])
            name = row['name']
            gender = row['gender']
            course_ids = row['courses'].split(',')
            course_list = [course for course in courses if course.code in course_ids]
            priority = row['priority'].strip().lower()
            faculty = Faculty(faculty_id, name, gender, course_list, priority)
            faculties.append(faculty)
    return faculties


def main():
    timetable = Timetable()

    courses = load_courses_from_csv('courses.csv')
    faculties = load_faculties_from_csv('faculty.csv', courses)
    slots_offered = ['L1+L2', 'L2+L3', 'L3+L4', 'L4+L5', 'L5+L6', 'L7+L8', 'L8+L9', 'L9+L10', 'L10+L11',
                       'L11+L12', 'L13+L14', 'L14+L15', 'L15+L16', 'L16+L17', 'L17+L18', 'L19+L20', 'L20+L21',
                       'L21+L22', 'L22+L23', 'L23+L24', 'L25+L26', 'L26+L27', 'L27+L28', 'L28+L29', 'L29+L30',
                       'L31+L32', 'L32+L33', 'L33+L34', 'L34+L35', 'L35+L36', 'L37+L38', 'L38+L39', 'L39+L40',
                       'L40+L41', 'L41+L42', 'L43+L44', 'L44+L45', 'L45+L46', 'L46+L47', 'L47+L48', 'L49+L50',
                       'L50+L51', 'L51+L52', 'L52+L53', 'L53+L54', 'L55+L56', 'L56+L57', 'L57+L58', 'L58+L59',
                       'L59+L60']
    rooms = ["CB-101", "AB2-456", "AB1-324", "CB-534"]

    # timetable.schedule_timetable_theory(faculties, rooms)
    timetable.schedule_research_scholar_lab_timetable(faculties, rooms)


main()
